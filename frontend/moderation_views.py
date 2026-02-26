"""
Moderation and Export Features
Admin moderation dashboard, verification system, and data export utilities
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse, FileResponse
from django.db.models import Q, Count
from django.utils import timezone
from django.conf import settings
from datetime import timedelta
from .models import Review, Company, ReviewFlag, DataExport, UserProfile
import re
import io
from pathlib import Path


# Spam keywords for detection
SPAM_KEYWORDS = [
    "casino",
    "viagra",
    "poker",
    "lottery",
    "win money",
    "click here",
    "buy now",
    "limited offer",
    "act now",
    "free money",
    "make money fast",
    "work from home",
    "weight loss",
    "miracle",
    "guaranteed",
]


def is_staff_user(user):
    """Check if user is staff/admin"""
    return user.is_staff or user.is_superuser


@staff_member_required
def moderation_dashboard(request):
    """Admin moderation dashboard for reviews"""
    # Get filter parameters
    status_filter = request.GET.get("status", "pending")
    search_query = request.GET.get("q", "")

    # Base queryset
    reviews = Review.objects.select_related("user", "company").all()

    # Apply filters
    if status_filter == "pending":
        reviews = reviews.filter(is_approved=False, approval_requested=True)
    elif status_filter == "approved":
        reviews = reviews.filter(is_approved=True)
    elif status_filter == "flagged":
        reviews = reviews.filter(flags__is_resolved=False).distinct()
    elif status_filter == "spam":
        # Reviews that might be spam
        spam_pattern = "|".join(SPAM_KEYWORDS)
        reviews = reviews.filter(
            Q(text__iregex=spam_pattern) | Q(title__iregex=spam_pattern)
        )

    # Search
    if search_query:
        reviews = reviews.filter(
            Q(text__icontains=search_query)
            | Q(user__username__icontains=search_query)
            | Q(company__name__icontains=search_query)
        )

    # Get flagged reviews
    flagged_reviews = (
        Review.objects.filter(flags__is_resolved=False)
        .distinct()
        .select_related("user", "company")
    )

    # Statistics
    stats = {
        "pending": Review.objects.filter(
            is_approved=False, approval_requested=True
        ).count(),
        "flagged": Review.objects.filter(flags__is_resolved=False).distinct().count(),
        "spam_detected": Review.objects.filter(
            Q(text__iregex="|".join(SPAM_KEYWORDS))
            | Q(title__iregex="|".join(SPAM_KEYWORDS))
        ).count(),
        "total_reviews": Review.objects.count(),
    }

    context = {
        "reviews": reviews[:100],  # Limit to 100 for performance
        "flagged_reviews": flagged_reviews[:20],
        "status_filter": status_filter,
        "search_query": search_query,
        "stats": stats,
    }

    return render(request, "frontend/moderation_dashboard.html", context)


@staff_member_required
def bulk_moderate_reviews(request):
    """Bulk approve/reject reviews"""
    if request.method == "POST":
        action = request.POST.get("action")
        review_ids = request.POST.getlist("review_ids[]")

        if not review_ids:
            return JsonResponse({"success": False, "error": "No reviews selected"})

        reviews = Review.objects.filter(id__in=review_ids)

        if action == "approve":
            company_ids = list(reviews.values_list("company_id", flat=True).distinct())
            approved_reviews = list(reviews.select_related("user", "company"))
            count = reviews.update(is_approved=True)
            from .utils import recalculate_company_stats
            from .email_notifications import EmailNotificationService
            for cid in company_ids:
                recalculate_company_stats(cid)
            for r in approved_reviews:
                try:
                    EmailNotificationService.send_review_approved_notification(r)
                except Exception:
                    pass
            return JsonResponse(
                {
                    "success": True,
                    "message": f"{count} ta sharh tasdiqlandi",
                    "count": count,
                }
            )

        elif action == "reject":
            company_ids = list(reviews.values_list("company_id", flat=True).distinct())
            rejected_reviews = list(reviews.select_related("user", "company"))
            count = reviews.count()
            from .utils import recalculate_company_stats
            from .email_notifications import EmailNotificationService
            for r in rejected_reviews:
                try:
                    EmailNotificationService.send_review_rejected_notification(r)
                except Exception:
                    pass
            reviews.delete()
            for cid in company_ids:
                recalculate_company_stats(cid)
            return JsonResponse(
                {
                    "success": True,
                    "message": f"{count} ta sharh o'chirildi",
                    "count": count,
                }
            )

        elif action == "spam":
            company_ids = list(reviews.values_list("company_id", flat=True).distinct())
            count = reviews.count()
            reviews.delete()
            from .utils import recalculate_company_stats
            for cid in company_ids:
                recalculate_company_stats(cid)
            return JsonResponse(
                {
                    "success": True,
                    "message": f"{count} ta spam sharh o'chirildi",
                    "count": count,
                }
            )

    return JsonResponse({"success": False, "error": "Invalid request"})


@login_required
def flag_review(request, review_id):
    """Flag a review for moderation"""
    review = get_object_or_404(Review, id=review_id)

    if request.method == "POST":
        reason = request.POST.get("reason", "other")
        description = request.POST.get("description", "")

        # Check if user already flagged this review
        existing_flag = ReviewFlag.objects.filter(
            review=review, flagged_by=request.user
        ).first()

        if existing_flag:
            return JsonResponse(
                {"success": False, "error": "Siz bu sharhni allaqachon belgiladingiz"}
            )

        # Create flag
        ReviewFlag.objects.create(
            review=review,
            flagged_by=request.user,
            reason=reason,
            description=description,
        )

        return JsonResponse(
            {"success": True, "message": "Sharh moderatsiyaga yuborildi"}
        )

    return JsonResponse({"success": False, "error": "Invalid request"})


@staff_member_required
def resolve_flag(request, flag_id):
    """Resolve a flagged review"""
    flag = get_object_or_404(ReviewFlag, id=flag_id)

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "delete":
            flag.review.delete()
            flag.is_resolved = True
            flag.action_taken = "Deleted review"
        elif action == "approve":
            flag.review.is_approved = True
            flag.review.save()
            flag.is_resolved = True
            flag.action_taken = "Approved review"
        elif action == "ignore":
            flag.is_resolved = True
            flag.action_taken = "Ignored flag"

        flag.resolved_by = request.user
        flag.resolved_at = timezone.now()
        flag.save()

        return JsonResponse({"success": True, "message": "Flag hal qilindi"})

    return JsonResponse({"success": False, "error": "Invalid request"})


# ============================================
# BUSINESS VERIFICATION
# ============================================


@login_required
def request_verification(request, company_id):
    """Business owner requests verification"""
    company = get_object_or_404(Company, id=company_id, manager=request.user)

    if request.method == "POST":
        if company.is_verified:
            return JsonResponse(
                {"success": False, "error": "Biznesingiz allaqachon tasdiqlangan"}
            )

        # Upload verification document
        document = request.FILES.get("document")
        if document:
            company.verification_document = document
            company.verification_requested_at = timezone.now()
            company.save()

            return JsonResponse(
                {
                    "success": True,
                    "message": "Tasdiqlash so'rovi yuborildi. Admin ko'rib chiqadi.",
                }
            )
        else:
            return JsonResponse(
                {"success": False, "error": "Tasdiqlash hujjatini yuklang"}
            )

    context = {"company": company}

    return render(request, "frontend/request_verification.html", context)


@staff_member_required
def approve_verification(request, company_id):
    """Admin approves business verification"""
    company = get_object_or_404(Company, id=company_id)

    if request.method == "POST":
        action = request.POST.get("action")
        notes = request.POST.get("notes", "")

        if action == "approve":
            company.is_verified = True
            company.verification_notes = notes
            company.save()

            return JsonResponse({"success": True, "message": "Biznes tasdiqlandi"})
        elif action == "reject":
            company.verification_document = None
            company.verification_requested_at = None
            company.verification_notes = notes
            company.save()

            return JsonResponse({"success": True, "message": "Tasdiqlash rad etildi"})

    # List pending verifications
    pending_verifications = Company.objects.filter(
        verification_requested_at__isnull=False, is_verified=False
    ).select_related("manager")

    context = {"company": company, "pending_verifications": pending_verifications}

    return render(request, "frontend/approve_verification.html", context)


# ============================================
# DATA EXPORT
# ============================================


@login_required
def export_reviews_pdf(request, company_id):
    """Export company reviews to PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        company = get_object_or_404(Company, id=company_id)
        reviews = (
            Review.objects.filter(company=company, is_approved=True)
            .select_related("user")
            .order_by("-created_at")
        )

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)

        # Container for elements
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#00d68b"),
            spaceAfter=30,
        )

        elements.append(Paragraph(f"Reviews - {company.name}", title_style))
        elements.append(Spacer(1, 0.2 * inch))

        # Company info
        info_style = styles["Normal"]
        elements.append(
            Paragraph(f"<b>Total Reviews:</b> {reviews.count()}", info_style)
        )
        elements.append(
            Paragraph(
                f"<b>Average Rating:</b> {company.rating}/5.0", info_style
            )
        )
        elements.append(
            Paragraph(
                f"<b>Generated:</b> {timezone.now().strftime('%Y-%m-%d %H:%M')}",
                info_style,
            )
        )
        elements.append(Spacer(1, 0.3 * inch))

        # Reviews table
        data = [["Date", "User", "Rating", "Review"]]

        for review in reviews[:100]:  # Limit to 100 reviews
            user_name = review.user.get_full_name() if review.user else "Anonymous"
            review_text = (
                review.text[:200] + "..." if len(review.text) > 200 else review.text
            )

            data.append(
                [
                    review.created_at.strftime("%Y-%m-%d"),
                    user_name,
                    f"{review.rating}/5",
                    review_text,
                ]
            )

        table = Table(data, colWidths=[1.2 * inch, 1.5 * inch, 0.8 * inch, 4 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )

        elements.append(table)

        # Build PDF
        doc.build(elements)

        # Return response
        buffer.seek(0)
        response = FileResponse(
            buffer, as_attachment=True, filename=f"{company.name}_reviews.pdf"
        )
        response["Content-Type"] = "application/pdf"

        return response

    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


@login_required
def export_reviews_excel(request, company_id):
    """Export company reviews to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        company = get_object_or_404(Company, id=company_id)
        reviews = (
            Review.objects.filter(company=company, is_approved=True)
            .select_related("user")
            .order_by("-created_at")
        )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reviews"

        # Headers
        headers = [
            "Date",
            "User",
            "Rating",
            "Title",
            "Review",
            "Helpful Votes",
            "Response",
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(
            start_color="00D68B", end_color="00D68B", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data
        for review in reviews:
            user_name = review.user.get_full_name() if review.user else "Anonymous"
            ws.append(
                [
                    review.created_at.strftime("%Y-%m-%d %H:%M"),
                    user_name,
                    review.rating,
                    review.title or "",
                    review.text,
                    review.helpful_count,
                    review.owner_response_text or "",
                ]
            )

        # Adjust column widths
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 50

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f"attachment; filename={company.name}_reviews.xlsx"
        )

        return response

    except Exception as e:
        return HttpResponse(f"Error generating Excel: {str(e)}", status=500)


@login_required
def export_user_data(request):
    """Export user's personal data (GDPR compliance)"""
    try:
        import json
        from openpyxl import Workbook

        user = request.user

        # Gather user data
        user_data = {
            "personal_info": {
                "username": user.username,
                "email": user.email,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "date_joined": user.date_joined.isoformat(),
                "last_login": user.last_login.isoformat() if user.last_login else None,
            },
            "profile": {},
            "reviews": [],
            "companies": [],
            "badges": [],
        }

        # Profile data
        if hasattr(user, "profile"):
            profile = user.profile
            user_data["profile"] = {
                "bio": profile.bio,
                "phone": profile.phone,
                "city": profile.city,
            }

        # Reviews
        reviews = Review.objects.filter(user=user)
        user_data["reviews"] = [
            {
                "company": r.company.name,
                "rating": r.rating,
                "title": r.title,
                "text": r.text,
                "date": r.created_at.isoformat(),
            }
            for r in reviews
        ]

        # Managed companies
        companies = Company.objects.filter(manager=user)
        user_data["companies"] = [
            {
                "name": c.name,
                "category": c.category_fk.name if c.category_fk else c.category,
                "rating": float(c.rating),
                "review_count": c.review_count,
            }
            for c in companies
        ]

        # Gamification badges
        if hasattr(user, "badges"):
            user_data["badges"] = [
                {
                    "name": b.name,
                    "description": b.description,
                    "earned_at": b.earned_at.isoformat(),
                }
                for b in user.badges.all()
            ]

        # Create JSON file
        buffer = io.BytesIO()
        json_str = json.dumps(user_data, indent=2, ensure_ascii=False)
        buffer.write(json_str.encode("utf-8"))
        buffer.seek(0)

        response = HttpResponse(buffer.read(), content_type="application/json")
        response["Content-Disposition"] = (
            f"attachment; filename={user.username}_data.json"
        )

        return response

    except Exception as e:
        return HttpResponse(f"Error exporting data: {str(e)}", status=500)


@login_required
def request_data_export(request):
    """Request async data export (for large datasets)"""
    if request.method == "POST":
        export_type = request.POST.get("export_type")
        company_id = request.POST.get("company_id")

        filters = {}
        if company_id:
            filters["company_id"] = company_id

        # Create export request
        export = DataExport.objects.create(
            user=request.user,
            export_type=export_type,
            filters=filters,
            expires_at=timezone.now() + timedelta(days=7),
        )

        # In production, trigger async task with Celery
        # For now, return success

        return JsonResponse(
            {
                "success": True,
                "message": "Export yaratilmoqda. Tayyor bo'lganda email yuboriladi.",
                "export_id": export.id,
            }
        )

    # Show export options
    exports = DataExport.objects.filter(user=request.user).order_by("-created_at")[:10]

    context = {"exports": exports}

    return render(request, "frontend/request_export.html", context)


# ---------------------------------------------------------------------------
# Telegram Bot Webhook — handles Approve / Reject button presses
# ---------------------------------------------------------------------------
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import json as _json


@csrf_exempt
@require_POST
def telegram_webhook(request):
    """Receive callback queries from Telegram inline keyboard buttons.

    Security: verified via X-Telegram-Bot-Api-Secret-Token header which
    Telegram sends with every webhook request when a secret_token is set
    during setWebhook. Falls back to checking TELEGRAM_BOT_TOKEN if not set.
    """
    from django.conf import settings
    from .models import Review
    from .utils import answer_telegram_callback, edit_telegram_message
    from .cache_utils import clear_public_cache

    token = getattr(settings, "TELEGRAM_BOT_TOKEN", "")
    if not token:
        return JsonResponse({"ok": False}, status=403)

    # Verify secret token header
    expected_secret = getattr(settings, "TELEGRAM_WEBHOOK_SECRET", token)
    incoming_secret = request.headers.get("X-Telegram-Bot-Api-Secret-Token", "")
    if incoming_secret != expected_secret:
        return JsonResponse({"ok": False}, status=403)

    try:
        body = _json.loads(request.body)
    except Exception:
        return JsonResponse({"ok": True})

    callback_query = body.get("callback_query")
    if not callback_query:
        return JsonResponse({"ok": True})

    cq_id = callback_query.get("id")
    chat_id = str(callback_query.get("message", {}).get("chat", {}).get("id", ""))
    message_id = callback_query.get("message", {}).get("message_id")
    data = callback_query.get("data", "")
    actor_name = callback_query.get("from", {}).get("first_name", "Admin")

    if ":" not in data:
        return JsonResponse({"ok": True})

    action, _, review_id_str = data.partition(":")
    if not review_id_str.isdigit():
        return JsonResponse({"ok": True})

    review_id = int(review_id_str)

    try:
        review = Review.objects.select_related("company", "user").get(pk=review_id)
    except Review.DoesNotExist:
        answer_telegram_callback(cq_id, "⚠️ Sharh topilmadi (allaqachon o'chirilgan bo'lishi mumkin).", token)
        if chat_id and message_id:
            edit_telegram_message(chat_id, message_id, "⚠️ <b>Sharh topilmadi</b> — allaqachon o'chirilgan.", token)
        return JsonResponse({"ok": True})

    from .email_notifications import EmailNotificationService

    if action == "approve":
        if review.is_approved:
            answer_telegram_callback(cq_id, "ℹ️ Bu sharh allaqachon tasdiqlangan.", token)
            return JsonResponse({"ok": True})
        review.is_approved = True
        review.approval_requested = True
        review.save(update_fields=["is_approved", "approval_requested"])
        # Recalculate company stats
        from .utils import recalculate_company_stats
        recalculate_company_stats(review.company_id)
        clear_public_cache()
        # Notify author
        try:
            EmailNotificationService.send_review_approved_notification(review)
        except Exception:
            pass
        answer_telegram_callback(cq_id, "✅ Sharh tasdiqlandi!", token)
        new_text = (
            f"✅ <b>Tasdiqlandi</b> — {actor_name}\n\n"
            f"🏢 {review.company.name}\n"
            f"👤 {review.user_name}  ⭐ {review.rating}/5\n"
            f"📝 {review.text[:300]}"
        )
        if chat_id and message_id:
            edit_telegram_message(chat_id, message_id, new_text, token)

    elif action == "reject":
        company_name = review.company.name
        user_name_str = review.user_name
        rating = review.rating
        # Notify author before deletion
        try:
            EmailNotificationService.send_review_rejected_notification(review)
        except Exception:
            pass
        review.delete()
        clear_public_cache()
        answer_telegram_callback(cq_id, "🗑 Sharh o'chirildi.", token)
        new_text = (
            f"❌ <b>O'chirildi</b> — {actor_name}\n\n"
            f"🏢 {company_name}\n"
            f"👤 {user_name_str}  ⭐ {rating}/5"
        )
        if chat_id and message_id:
            edit_telegram_message(chat_id, message_id, new_text, token)

    return JsonResponse({"ok": True})

